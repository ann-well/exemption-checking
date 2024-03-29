import tkinter as tk
from tkinter import ttk
import datetime, os, sys
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import openpyxl
from openpyxl.styles import PatternFill
import send2trash
from tkinter import messagebox

def resource_path(relative_path):
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# Create window for simple GUI
window = tk.Tk()
window.title('Exemption reason checker')
photo = tk.PhotoImage(file=resource_path('exemptionLogo.png'))
image_label = ttk.Label(
    window,
    image=photo,
    padding=5
)
image_label.pack()

# Open daily report file
try:
    wb = openpyxl.load_workbook('Expired Exemption with Email.xlsx')
    sheet = wb['Page1_1']

    # Determine the output file name from the date of the report
    reportDate = sheet.cell(row=sheet.max_row, column=1).value
    fileName = str(reportDate.day).zfill(2) + str(reportDate.month).zfill(2) + str(reportDate.year)
except:  # No file error handling
    messagebox.showwarning("No file detected",
                           "No file Expired Exemption with Email.xlsx found.\nPlease exit and put the file in this folder")


# Function to run from tk inter window
def checkExemption():
    day = []
    for i in range(3, sheet.max_row):
        day.append(sheet.cell(row=i, column=1).value)
    month = []
    for i in range(3, sheet.max_row):
        month.append(sheet.cell(row=i, column=2).value)
    year = []
    for i in range(3, sheet.max_row):
        year.append(sheet.cell(row=i, column=3).value)
    name = []
    for i in range(3, sheet.max_row):
        name.append(sheet.cell(row=i, column=4).value)
    lastname = []
    for i in range(3, sheet.max_row):
        lastname.append(sheet.cell(row=i, column=5).value)
    code = []
    for i in range(3, sheet.max_row):
        code.append(sheet.cell(row=i, column=6).value)
    siebel = []
    for i in range(3, sheet.max_row):
        siebel.append(sheet.cell(row=i, column=7).value)

    # Add column for output dates
    sheet.insert_cols(8)  # add a column for expiry dates
    sheet.column_dimensions['H'].width = 30

    # Create pattern fills
    fillRed = PatternFill(patternType='solid', fgColor='FFC7CE')  # create pattern fills
    fillGreen = PatternFill(patternType='solid', fgColor='C6EFCE')

    # Open NHS site
    browser = webdriver.Edge(service=Service(EdgeChromiumDriverManager().install()))
    browser.get("https://services.nhsbsa.nhs.uk/check-my-nhs-exemption/start")

    for i in range(len(day)):  # Loop for the number of patients in report
        # Find and populate form elements with data from report
        try:
            first = browser.find_element('id', 'next-button')
            first.submit()
            time.sleep(0.5)
            dayWeb = browser.find_element('id', 'dob-day')
            dayWeb.send_keys(day[i])

            monthWeb = browser.find_element('id', 'dob-month')
            monthWeb.send_keys(month[i])

            yearWeb = browser.find_element('id', 'dob-year')
            yearWeb.send_keys(year[i])
            yearWeb.submit()
            time.sleep(0.5)

            nameWeb = browser.find_element('id', 'firstname')
            nameWeb.send_keys(name[i])

            lastnameWeb = browser.find_element('id', 'lastname')
            lastnameWeb.send_keys(lastname[i])
            lastnameWeb.submit()
            time.sleep(0.5)

            codeWeb = browser.find_element('id', 'postcode')
            try:
                codeWeb.send_keys(code[i])
                codeWeb.submit()
                time.sleep(0.5)
            except:  # No postcode error handling
                browser.get("https://services.nhsbsa.nhs.uk/check-my-nhs-exemption/start")
                sheet.cell(row=i + 3, column=8).value = 'No postcode'
                for y in range(1, 17):
                    sheet.cell(row=i + 3, column=y).fill = fillRed
                continue
        except:  # Merged/duplicate cells error handling
            if sheet.cell(row=i + 2, column=y).fill == fillRed:
                for y in range(1, 17):
                    sheet.cell(row=i + 3, column=y).fill = fillRed
            browser.get("https://services.nhsbsa.nhs.uk/check-my-nhs-exemption/start")
            continue
        
        # Result of submitting the form
        result = browser.find_element(By.CSS_SELECTOR, '.nhsuk-heading-xl').text  # webbrowser result

        # NHS exemption present
        if result == 'You currently have an NHS exemption':
            exemptionDate = browser.find_element(By.CSS_SELECTOR, '.exemption-done-panel > h2:nth-child(2)').text
            empty, expireText, date = exemptionDate.partition('Expires on ')  # convert date to siebel-ready format
            actualDate = datetime.datetime.strptime(date, '%d %B %Y')
            sheet.cell(row=i + 3, column=8).value = actualDate.strftime('%d.%m.%Y')
            for y in range(1, 17):
                sheet.cell(row=i + 3, column=y).fill = fillGreen

        # Over 60 years old
        elif result == 'You get help with health costs':
            sheet.cell(row=i + 3, column=8).value = '60 years old'
            for y in range(1, 17):
                sheet.cell(row=i + 3, column=y).fill = fillGreen

        # No information about the patient
        elif result == '''We couldn't match you to our records''':
            for y in range(1, 17):
                sheet.cell(row=i + 3, column=y).fill = fillRed

        end = browser.find_element(By.CSS_SELECTOR, '.nhsuk-action-link__text')  # back to front page
        end.click()
        time.sleep(1)

    # Save created file
    wb.save(fileName + '.xlsx')

    # Delete input file
    send2trash.send2trash('Expired Exemption with Email.xlsx')

    # Close windows
    window.destroy()


# Start the function by user input
addButton = tk.Button(window,
                      text="START",
                      command=checkExemption)
addButton.pack()

window.mainloop()

