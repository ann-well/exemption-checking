from selenium import webdriver
from selenium.webdriver.edge.service import Service
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import openpyxl
from openpyxl.styles import PatternFill
import datetime

# Timer to check the runtime
startTime = time.time()

# Open the file (--change the file name in the code--)
wb = openpyxl.load_workbook('Unknown Dec to do .xlsx')

# Function to check specific sheet from the report
def checkSheet(sheetName):
    sheet = wb[sheetName]
    # Var to count the updated records
    allUpdated = 0

    # Create lists of data from the report
    day = []
    for i in range(2, sheet.max_row + 1):
        day.append(sheet.cell(row= i, column= 1).value)
    month = []
    for i in range(2, sheet.max_row + 1):
        month.append(sheet.cell(row= i, column= 2).value)
    year = []
    for i in range(2, sheet.max_row + 1):
        year.append(sheet.cell(row= i, column= 3).value)
    name = []
    for i in range(2, sheet.max_row + 1):
        name.append(sheet.cell(row= i, column= 4).value)
    lastname = []
    for i in range(2, sheet.max_row + 1):
        lastname.append(sheet.cell(row= i, column= 5).value)
    code = []
    for i in range(2, sheet.max_row + 1):
        code.append(sheet.cell(row= i, column= 6).value)
    siebel = []
    for i in range(2, sheet.max_row + 1):
        siebel.append(sheet.cell(row= i, column= 7).value)

    sheet.insert_cols(8)
    sheet.column_dimensions['H'].width = 30

    # Create pattern fills to later use in output file
    fillRed = PatternFill(patternType='solid', fgColor='FFC7CE')
    fillYel = PatternFill(patternType='solid', fgColor='FFEB9C')
    fillGreen = PatternFill(patternType='solid', fgColor='C6EFCE')

    # Open NHS webpage with the form
    browser = webdriver.Edge(service=Service(EdgeChromiumDriverManager().install()))
    browser.get("https://services.nhsbsa.nhs.uk/check-my-nhs-exemption/start")

    # Populate the form with the data from report
    for i in range(len(day)):
        try:
            first = browser.find_element('id', 'next-button')
            first.submit()
            time.sleep(1)
            dayWeb = browser.find_element('id', 'dob-day')
            dayWeb.send_keys(day[i])

            monthWeb = browser.find_element('id', 'dob-month')
            monthWeb.send_keys(month[i])

            yearWeb = browser.find_element('id', 'dob-year')
            yearWeb.send_keys(year[i])
            yearWeb.submit()
            time.sleep(0.5)
            try:
                # Exception: patient under 16 years old
                age = browser.find_element(By.CSS_SELECTOR, '.nhsuk-heading-xl').text
                if age == 'You get help with health costs':
                    print(name[i], lastname[i], 'under 16 years old')
                    sheet.cell(row=i+2, column =8).value = 'under 16 years old'
                    allUpdated += 1
                    end = browser.find_element(By.CSS_SELECTOR, '.nhsuk-action-link__text')
                    end.click()
                    time.sleep(1)
                    continue
            except:
                nameWeb = browser.find_element('id', 'firstname')
                nameWeb.send_keys(name[i])

                lastnameWeb = browser.find_element('id', 'lastname')
                lastnameWeb.send_keys(lastname[i])
                lastnameWeb.submit()
                time.sleep(0.5)

                codeWeb = browser.find_element('id', 'postcode')
                codeWeb.send_keys(code[i])
                codeWeb.submit()
                time.sleep(1)

                
                result = browser.find_element(By.CSS_SELECTOR, '.nhsuk-heading-xl').text
                print(result)
                # Exemption present - insert the end date into output file
                if result == 'You currently have an NHS exemption':
                    exemptionDate = browser.find_element(By.CSS_SELECTOR, '.exemption-done-panel > h2:nth-child(2)').text
                    empty, expireText, date = exemptionDate.partition('Expires on ')  # convert date to siebel-ready format
                    actualDate = datetime.datetime.strptime(date, '%d %B %Y')
                    sheet.cell(row=i + 2, column=8).value = actualDate.strftime('%d.%m.%Y')
                    for y in range(1, 17):
                        sheet.cell(row=i + 2, column=y).fill = fillGreen
                    allUpdated += 1

                # Patient over 60 years old
                elif result == 'You get help with health costs':
                    print(name[i], lastname[i], '60 years old')
                    sheet.cell(row=i+2, column =8).value = '60 years old'
                    for y in range(1, 17):
                        sheet.cell(row=i + 2, column=y).fill = fillGreen
                    allUpdated += 1

                # No exemption found
                elif result == '''We couldn't match you to our records''':
                    for y in range(1, 17):
                        sheet.cell(row=i+2, column=y).fill = fillRed

            end = browser.find_element(By.CSS_SELECTOR, '.nhsuk-action-link__text')
            end.click()
            time.sleep(1)
        # All other exceptions - no postcodes, wrong format in names etc
        except:
            browser.get("https://services.nhsbsa.nhs.uk/check-my-nhs-exemption/start")
            for y in range(1, 17):
                sheet.cell(row=i+2, column=y).fill = fillYel
            continue

    # Insert the number of all records and updated records into output file
    sheet.cell(row =sheet.max_row + 2, column = 1).value = 'All:'
    sheet.cell(row=sheet.max_row, column=2).value = len(day)
    sheet.cell(row=sheet.max_row, column=4).value = 'Updated:'
    sheet.cell(row=sheet.max_row, column=5).value = allUpdated

# Check both sheets (--update the code if different--)
checkSheet('England')
checkSheet('Sheet2')


# Runtime check
endTime = time.time()
ile = endTime - startTime
print('TIME:')
print(round(ile, 2))

# Save the output file
wb.save('prechecked Unknowns to do.xlsx')
