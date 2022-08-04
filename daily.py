from selenium import webdriver
from selenium.webdriver.edge.service import Service
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook('test.xlsx')
sheet = wb['Charter']

day = []
for i in range(3, sheet.max_row):
    day.append(sheet.cell(row= i, column= 1).value)  
month = []
for i in range(3, sheet.max_row):
    month.append(sheet.cell(row= i, column= 2).value)
year = []
for i in range(3, sheet.max_row):
    year.append(sheet.cell(row= i, column= 3).value)
name = []
for i in range(3, sheet.max_row):
    name.append(sheet.cell(row= i, column= 4).value)
lastname = []
for i in range(3, sheet.max_row):
    lastname.append(sheet.cell(row= i, column= 5).value)
code = []
for i in range(3, sheet.max_row):
    code.append(sheet.cell(row= i, column= 6).value)
siebel = []
for i in range(3, sheet.max_row):
    siebel.append(sheet.cell(row= i, column= 7).value)

sheet.insert_cols(8)
sheet.column_dimensions['H'].width = 30
fillRed = PatternFill(patternType='solid', fgColor='FFC7CE')

browser = webdriver.Edge(service=Service(EdgeChromiumDriverManager().install()))
browser.get(https://services.nhsbsa.nhs.uk/check-my-nhs-exemption/start)

for i in range(len(day)):
    try:
        first = browser.find_element('id', 'next-button')
        first.submit()
        time.sleep(0.5)
        dayWeb = browser.find_element('id', 'dob-day')
        dayWeb.send_keys(day[i])

        monthWeb = browser.find_element('id', 'dob-month')
        monthWeb.send_keys(month[i])

        yearWeb = browser.find_element('id', 'dob-year')
        yearWeb.send_keys(year[i])
        yearWeb.submit()

        time.sleep(0.5)

        nameWeb = browser.find_element('id', 'firstname')
        nameWeb.send_keys(name[i])

        lastnameWeb = browser.find_element('id', 'lastname')
        lastnameWeb.send_keys(lastname[i])
        lastnameWeb.submit()
        time.sleep(0.5)

        codeWeb = browser.find_element('id', 'postcode')
        try:
            codeWeb.send_keys(code[i])
            codeWeb.submit()
            time.sleep(0.5)
        except:
            browser.get(https://services.nhsbsa.nhs.uk/check-my-nhs-exemption/start)
            sheet.cell(row=i+3, column =8).value = 'No postcode'
            continue
    except:
        if sheet.cell(row=i+2, column=y).fill == fillRed:
            for y in range(1, 17):
                sheet.cell(row=i+3, column=y).fill = fillRed
        browser.get(https://services.nhsbsa.nhs.uk/check-my-nhs-exemption/start)
        continue

    result = browser.find_element(By.CSS_SELECTOR, '.nhsuk-heading-xl').text
    if result == 'You currently have an NHS exemption':
        exemptionDate = browser.find_element(By.CSS_SELECTOR,'.exemption-done-panel > h2:nth-child(2)').text
        sheet.cell(row=i+3, column =8).value = exemptionDate
    elif result == 'You get help with health costs':
        sheet.cell(row=i+3, column =8).value = '60 years old'

    elif result == '''We couldn't match you to our records''':
        for y in range(1, 17):
            sheet.cell(row=i+3, column=y).fill = fillRed

    end = browser.find_element(By.CSS_SELECTOR, '.nhsuk-action-link__text')
    end.click()
    time.sleep(1)

wb.save('testing.xlsx')
